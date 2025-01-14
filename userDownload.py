import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import requests
import zipfile
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import re


def get_column_width(text):  # """열 너비 계산 (한글 2칸, 영문 1칸)"""
    if not text:
        return 0
    length = 0
    for c in str(text):
        length += 2 if ord(c) > 127 else 1
    return length


def process_excel(df, excel_buffer):  # """엑셀 파일을 메모리에 생성하고 스타일 적용"""
    df_excel = df.copy()
    df_excel.index = range(1, len(df) + 1)

    # URL-파일명 쌍 찾기
    url_filename_pairs = []
    for i in range(1, 11):
        url_col = f"첨부파일 URL ({i})"
        filename_col = f"첨부파일명 ({i})"
        if url_col in df.columns and filename_col in df.columns:
            url_filename_pairs.append((url_col, filename_col))

    # 파일명 컬럼 삭제
    filename_cols = [pair[1] for pair in url_filename_pairs]
    df_excel = df_excel.drop(columns=filename_cols)

    # 엑셀 파일 생성
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=True, index_label="No")
        wb = writer.book
        ws = wb.active

        # 스타일 설정
        url_style = Font(color="0563C1", underline="single")
        center_align = Alignment(horizontal="center", vertical="center")

        # 열 너비 계산용 딕셔너리
        column_widths = {"A": get_column_width("No")}

        # 헤더 처리
        for col_idx, cell in enumerate(ws[1], 1):
            column_widths[get_column_letter(col_idx)] = max(
                get_column_width(str(cell.value)),
                column_widths.get(get_column_letter(col_idx), 0),
            )
            cell.alignment = center_align

        # 데이터 처리
        for row_idx, row in enumerate(list(ws.rows)[1:], 1):
            for col_idx, cell in enumerate(row):
                col_name = ws.cell(row=1, column=cell.column).value

                # URL 컬럼 처리
                for url_col, filename_col in url_filename_pairs:
                    if col_name == url_col and pd.notna(cell.value):
                        filename = df.iloc[row_idx - 1][filename_col]
                        cell.value = filename
                        cell.hyperlink = df.iloc[row_idx - 1][url_col]
                        cell.font = url_style
                        break

                # 열 너비 업데이트
                width = get_column_width(str(cell.value))
                col_letter = get_column_letter(cell.column)
                if width > column_widths.get(col_letter, 0):
                    column_widths[col_letter] = width

                cell.alignment = center_align

        # 열 너비 설정
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width + 4


def create_download_buttons(df, category):
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    col1, col2 = st.columns([1, 2])

    with col1:
        # 엑셀 다운로드 - 기존 코드 유지
        excel_buffer = BytesIO()
        process_excel(df, excel_buffer)

        excel_filename = f"입찰정보_{category}_{now}.xlsx"

        st.download_button(
            label="📊 Excel 다운로드",
            data=excel_buffer.getvalue(),
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col2:
        # 첨부파일 다운로드 로직 개선
        if st.button("📁 첨부파일 일괄 다운로드", key=f"{category}_files"):
            with st.spinner("첨부파일 다운로드 준비 중..."):
                zip_buffer = BytesIO()
                success_count = 0
                error_count = 0

                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    # 세션 재사용을 위한 설정
                    session = requests.Session()
                    session.mount(
                        "https://", requests.adapters.HTTPAdapter(max_retries=3)
                    )

                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    for idx, row in df.iterrows():
                        progress = idx / len(df)
                        progress_bar.progress(progress)

                        # 폴더명 생성
                        folder_name = re.sub(r'[<>:"/\\|?*]', "_", str(row["공고명"]))
                        folder_name = f"{idx}. {folder_name}"
                        status_text.text(f"처리 중: {folder_name}")

                        # 첨부파일 처리
                        for i in range(1, 11):
                            url_col = f"첨부파일 URL ({i})"
                            filename_col = f"첨부파일명 ({i})"

                            if url_col in df.columns and filename_col in df.columns:
                                if pd.notna(row[url_col]) and pd.notna(
                                    row[filename_col]
                                ):
                                    try:
                                        # 타임아웃 설정과 함께 요청
                                        response = session.get(row[url_col], timeout=10)
                                        if response.status_code == 200:
                                            filename = f"{i}) {row[filename_col]}"
                                            zip_path = f"{folder_name}/{filename}"
                                            zip_file.writestr(
                                                zip_path, response.content
                                            )
                                            success_count += 1
                                        else:
                                            error_count += 1
                                    except requests.Timeout:
                                        error_count += 1
                                        continue
                                    except Exception as e:
                                        error_count += 1
                                        continue

                    progress_bar.progress(1.0)
                    status_text.text("다운로드 완료!")

                    # 결과 표시
                    st.success(
                        f"다운로드 완료 (성공: {success_count}개, 실패: {error_count}개)"
                    )

                    # ZIP 파일 다운로드 버튼
                    st.download_button(
                        label="💾 ZIP 파일 다운로드",
                        data=zip_buffer.getvalue(),
                        file_name=f"첨부파일_{category}_{now}.zip",
                        mime="application/zip",
                        key=f"zip_{category}_{now}",
                    )

        # 첨부파일 목록 표시 개선
        with st.expander(":green[🔍 첨부파일 목록 보기]"):
            for idx, row in df.iterrows():
                folder_name = re.sub(r'[<>:"/\\|?*]', "_", str(row["공고명"]))
                folder_name = f"{idx}. {folder_name}"

                st.write(f"📁 {folder_name}")

                files_found = False
                for i in range(1, 11):
                    url_col = f"첨부파일 URL ({i})"
                    filename_col = f"첨부파일명 ({i})"
                    if url_col in df.columns and filename_col in df.columns:
                        if pd.notna(row[url_col]) and pd.notna(row[filename_col]):
                            files_found = True
                            filename = f"{i}) {row[filename_col]}"
                            col1, col2 = st.columns([3, 1])
                            col1.write(f"&nbsp; &nbsp; &nbsp; &nbsp; 📄 {filename}")
                            # 개별 다운로드 버튼은 URL만 표시
                            col2.markdown(f"[💾]({row[url_col]})")

                if not files_found:
                    st.write("&nbsp; &nbsp; &nbsp; &nbsp; (첨부파일 없음)")
